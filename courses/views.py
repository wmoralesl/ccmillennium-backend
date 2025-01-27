
from rest_framework import viewsets, status
from .models import Course, Group, Module, Enrollment
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.views import APIView
from rest_framework.response import Response
from .serializers import CourseSerializer, GroupSerializer, ModuleSerializer, EnrollmentSerializer
from django.template.loader import render_to_string
from rest_framework.decorators import api_view
from weasyprint import HTML
from django.http import HttpResponse
import openpyxl
from openpyxl.utils import get_column_letter

class CourseViewSet(viewsets.ModelViewSet):
    queryset = Course.objects.filter(is_active=True)
    serializer_class = CourseSerializer

class CourseGroupCountView(APIView):
    def get(self, request, course_id):
        try:
            course = Course.objects.get(id=course_id)
            group_count = course.group_count()
            return Response({'course': course.name, 'group_count': group_count})
        except Course.DoesNotExist:
            return Response({'error': 'Course not found'}, status=404)

class GroupViewSet(viewsets.ModelViewSet):
    queryset = Group.objects.all()
    serializer_class = GroupSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['teacher']

class ModuleViewSet(viewsets.ModelViewSet):
    queryset = Module.objects.all()
    serializer_class = ModuleSerializer

class ModulesByGroupView(APIView):
    def get(self, request, group_id):
        try:
            group = Group.objects.get(id=group_id)
            course = group.course
            if course:
                modules = Module.objects.filter(course=course)
                module_data = [{'id': module.id, 'name': module.name} for module in modules]
                return Response(module_data, status=status.HTTP_200_OK)
            else:
                return Response({"error": "Course not found for this group"}, status=status.HTTP_404_NOT_FOUND)
        except Group.DoesNotExist:
            return Response({"error": "Group not found"}, status=status.HTTP_404_NOT_FOUND)

class EnrollmentViewSet(viewsets.ModelViewSet):
    queryset = Enrollment.objects.all()
    serializer_class = EnrollmentSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['group', 'student']

    def create(self, request, *args, **kwargs):
        student = request.data.get('student')
        group = request.data.get('group')

        # Verifica si ya existe una inscripción para el estudiante y grupo
        if Enrollment.objects.filter(student=student, group=group).exists():
            return Response(
                {"detail": "El estudiante ya está asignado a este grupo."},
                status=status.HTTP_400_BAD_REQUEST
            )

        return super().create(request, *args, **kwargs)

@api_view(['GET'])
def generate_pdf(request, id):
    try:
        group = Group.objects.get(id=id)
        course_name = group.course.name
        enrollments = Enrollment.objects.filter(group=group)
        
        # Obtener detalles de los estudiantes
        student_details = []
        for enrollment in enrollments:
            student = enrollment.student
            student_details.append({
                'full_name': student.get_full_name(),
                'username': student.username,
                'dpi': student.dpi,
            })

        # Renderizar el template con los datos del grupo y estudiantes
        html_string = render_to_string('group_template.html', {
            'course_name': course_name,
            'student_details': student_details,
        })
        
        # Crear el PDF
        html = HTML(string=html_string)
        pdf = html.write_pdf()

        # Crear la respuesta
        response = HttpResponse(pdf, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="grupo_{id}.pdf"'
        return response

    except Group.DoesNotExist:
        return HttpResponse(status=404)
    
@api_view(['GET'])
def generate_xlsx(request, id):
    try:
        group = Group.objects.get(id=id)
        course_name = group.course.name
        enrollments = Enrollment.objects.filter(group=group)
        
        # Crear un nuevo libro de trabajo y una hoja
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = f"Grupo {id}"

        # Agregar encabezados
        headers = ["Full Name", "Username", "DPI"]
        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            sheet[f"{col_letter}1"] = header
            sheet[f"{col_letter}1"].font = openpyxl.styles.Font(bold=True)
            sheet.column_dimensions[col_letter].width = 20

        # Agregar detalles de los estudiantes
        for row_num, enrollment in enumerate(enrollments, start=2):
            student = enrollment.student
            sheet[f"A{row_num}"] = student.get_full_name()
            sheet[f"B{row_num}"] = student.username
            sheet[f"C{row_num}"] = student.dpi

        # Crear la respuesta
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="grupo_{id}.xlsx"'

        # Guardar el archivo en la respuesta
        workbook.save(response)
        return response

    except Group.DoesNotExist:
        return HttpResponse(status=404)